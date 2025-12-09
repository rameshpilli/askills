#!/usr/bin/env python3
"""
Script to generate sample PDF and Excel files for testing skills.
Run this script to create test data files.
"""

import os
from pathlib import Path

# Try to import required libraries
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    print("reportlab not installed. Run: pip install reportlab")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not installed. Run: pip install openpyxl")


def create_earnings_pdf(output_path: str):
    """Create a sample Q3 2024 Earnings Report PDF."""
    if not HAS_REPORTLAB:
        print("Skipping PDF creation - reportlab not installed")
        return

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("TechCorp Inc.", title_style))
    story.append(Paragraph("Q3 2024 Earnings Report", styles['Heading2']))
    story.append(Spacer(1, 20))

    # Executive Summary
    story.append(Paragraph("Executive Summary", styles['Heading3']))
    story.append(Paragraph(
        "TechCorp Inc. reported strong results for Q3 2024, with revenue increasing 15% "
        "year-over-year to $2.4 billion. Net income rose 22% to $340 million, driven by "
        "growth in our cloud services division and operational efficiency improvements.",
        styles['Normal']
    ))
    story.append(Spacer(1, 15))

    # Financial Highlights Table
    story.append(Paragraph("Financial Highlights (in millions USD)", styles['Heading3']))

    data = [
        ['Metric', 'Q3 2024', 'Q3 2023', 'Change'],
        ['Revenue', '$2,400', '$2,087', '+15%'],
        ['Gross Profit', '$1,680', '$1,418', '+18%'],
        ['Operating Income', '$480', '$376', '+28%'],
        ['Net Income', '$340', '$279', '+22%'],
        ['EPS (Diluted)', '$1.85', '$1.52', '+22%'],
    ]

    table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))

    # Segment Performance
    story.append(Paragraph("Segment Performance", styles['Heading3']))

    segments = [
        ['Segment', 'Revenue', '% of Total', 'YoY Growth'],
        ['Cloud Services', '$1,200M', '50%', '+25%'],
        ['Enterprise Software', '$720M', '30%', '+10%'],
        ['Hardware', '$360M', '15%', '+5%'],
        ['Professional Services', '$120M', '5%', '+8%'],
    ]

    seg_table = Table(segments, colWidths=[2*inch, 1.3*inch, 1.2*inch, 1.2*inch])
    seg_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
    ]))
    story.append(seg_table)
    story.append(Spacer(1, 20))

    # Outlook
    story.append(Paragraph("Q4 2024 Outlook", styles['Heading3']))
    story.append(Paragraph(
        "For Q4 2024, we expect revenue in the range of $2.5B to $2.6B, representing "
        "12-16% year-over-year growth. We anticipate continued strength in cloud services "
        "and are investing in AI capabilities to drive future growth.",
        styles['Normal']
    ))

    doc.build(story)
    print(f"Created: {output_path}")


def create_pricing_pdf(output_path: str):
    """Create a sample Product Pricing PDF."""
    if not HAS_REPORTLAB:
        print("Skipping PDF creation - reportlab not installed")
        return

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        spaceAfter=20,
        alignment=1
    )
    story.append(Paragraph("CloudPlatform Pro", title_style))
    story.append(Paragraph("Enterprise Pricing Guide 2024", styles['Heading2']))
    story.append(Spacer(1, 15))

    # Introduction
    story.append(Paragraph(
        "CloudPlatform Pro offers flexible pricing to meet your organization's needs. "
        "Choose from our tiered plans or contact us for custom enterprise solutions.",
        styles['Normal']
    ))
    story.append(Spacer(1, 20))

    # Pricing Tiers
    story.append(Paragraph("Subscription Tiers", styles['Heading3']))

    tiers = [
        ['Plan', 'Monthly', 'Annual', 'Users', 'Storage', 'Support'],
        ['Starter', '$29', '$290', 'Up to 5', '100 GB', 'Email'],
        ['Professional', '$99', '$990', 'Up to 25', '1 TB', 'Priority'],
        ['Business', '$299', '$2,990', 'Up to 100', '10 TB', '24/7 Phone'],
        ['Enterprise', 'Custom', 'Custom', 'Unlimited', 'Unlimited', 'Dedicated'],
    ]

    tier_table = Table(tiers, colWidths=[1.2*inch, 0.9*inch, 0.9*inch, 1*inch, 1*inch, 1*inch])
    tier_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3182ce')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ebf8ff')),
    ]))
    story.append(tier_table)
    story.append(Spacer(1, 20))

    # Add-ons
    story.append(Paragraph("Optional Add-ons", styles['Heading3']))

    addons = [
        ['Add-on', 'Description', 'Price/Month'],
        ['Extra Storage', 'Additional 1 TB block', '$20'],
        ['API Access', 'REST API with 10K calls/day', '$50'],
        ['SSO Integration', 'SAML/OIDC support', '$25'],
        ['Advanced Analytics', 'BI dashboards & reports', '$75'],
        ['Data Export', 'Automated backups to S3/GCS', '$30'],
    ]

    addon_table = Table(addons, colWidths=[1.5*inch, 3*inch, 1.2*inch])
    addon_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#48bb78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
    ]))
    story.append(addon_table)
    story.append(Spacer(1, 20))

    # Volume Discounts
    story.append(Paragraph("Volume Discounts", styles['Heading3']))
    story.append(Paragraph(
        "• 10-49 licenses: 10% discount<br/>"
        "• 50-99 licenses: 15% discount<br/>"
        "• 100-499 licenses: 20% discount<br/>"
        "• 500+ licenses: Contact sales for custom pricing",
        styles['Normal']
    ))
    story.append(Spacer(1, 15))

    # Contact
    story.append(Paragraph("Contact Sales", styles['Heading3']))
    story.append(Paragraph(
        "Email: sales@cloudplatform.example.com<br/>"
        "Phone: 1-800-CLOUD-PRO<br/>"
        "Web: www.cloudplatform.example.com/enterprise",
        styles['Normal']
    ))

    doc.build(story)
    print(f"Created: {output_path}")


def create_pricing_excel(output_path: str):
    """Create a sample pricing spreadsheet."""
    if not HAS_OPENPYXL:
        print("Skipping Excel creation - openpyxl not installed")
        return

    wb = Workbook()

    # Sheet 1: Pricing Tiers
    ws1 = wb.active
    ws1.title = "Pricing Tiers"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3182CE", end_color="3182CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Plan', 'Monthly Price', 'Annual Price', 'Max Users', 'Storage (GB)', 'API Calls/Day', 'Support Level']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data
    data = [
        ['Free', 0, 0, 1, 5, 100, 'Community'],
        ['Starter', 29, 290, 5, 100, 1000, 'Email'],
        ['Professional', 99, 990, 25, 1000, 10000, 'Priority'],
        ['Business', 299, 2990, 100, 10000, 100000, '24/7 Phone'],
        ['Enterprise', 999, 9990, 500, 100000, 1000000, 'Dedicated'],
    ]

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [2, 3]:  # Price columns
                cell.number_format = '$#,##0'
            elif col_num in [5, 6]:  # Number columns
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='center')

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 2: Revenue Calculator
    ws2 = wb.create_sheet("Revenue Calculator")

    ws2['A1'] = 'Revenue Calculator'
    ws2['A1'].font = Font(bold=True, size=14)

    calc_headers = ['Plan', 'Price', 'Customers', 'Monthly Revenue', 'Annual Revenue']
    for col, header in enumerate(calc_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    plans = [
        ['Free', 0, 1000],
        ['Starter', 29, 500],
        ['Professional', 99, 200],
        ['Business', 299, 50],
        ['Enterprise', 999, 10],
    ]

    for row_num, (plan, price, customers) in enumerate(plans, 4):
        ws2.cell(row=row_num, column=1, value=plan).border = border
        ws2.cell(row=row_num, column=2, value=price).border = border
        ws2.cell(row=row_num, column=2).number_format = '$#,##0'
        ws2.cell(row=row_num, column=3, value=customers).border = border

        # Monthly Revenue formula
        monthly_cell = ws2.cell(row=row_num, column=4)
        monthly_cell.value = f'=B{row_num}*C{row_num}'
        monthly_cell.number_format = '$#,##0'
        monthly_cell.border = border

        # Annual Revenue formula
        annual_cell = ws2.cell(row=row_num, column=5)
        annual_cell.value = f'=D{row_num}*12'
        annual_cell.number_format = '$#,##0'
        annual_cell.border = border

    # Totals
    total_row = 4 + len(plans)
    ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})').font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=f'=SUM(D4:D{total_row-1})').font = Font(bold=True)
    ws2.cell(row=total_row, column=4).number_format = '$#,##0'
    ws2.cell(row=total_row, column=5, value=f'=SUM(E4:E{total_row-1})').font = Font(bold=True)
    ws2.cell(row=total_row, column=5).number_format = '$#,##0'

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"Created: {output_path}")


def create_earnings_excel(output_path: str):
    """Create a sample earnings spreadsheet."""
    if not HAS_OPENPYXL:
        print("Skipping Excel creation - openpyxl not installed")
        return

    wb = Workbook()

    # Sheet 1: Quarterly Results
    ws1 = wb.active
    ws1.title = "Quarterly Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws1['A1'] = 'TechCorp Inc. - Financial Summary 2024'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:F1')

    headers = ['Metric', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024 (Est)', 'FY 2024 (Est)']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    data = [
        ['Revenue ($M)', 2100, 2250, 2400, 2550, '=SUM(B4:E4)'],
        ['Cost of Revenue ($M)', 735, 765, 720, 790, '=SUM(B5:E5)'],
        ['Gross Profit ($M)', '=B4-B5', '=C4-C5', '=D4-D5', '=E4-E5', '=SUM(B6:E6)'],
        ['Operating Expenses ($M)', 980, 1020, 1200, 1100, '=SUM(B7:E7)'],
        ['Operating Income ($M)', '=B6-B7', '=C6-C7', '=D6-D7', '=E6-E7', '=SUM(B8:E8)'],
        ['Net Income ($M)', 280, 310, 340, 370, '=SUM(B9:E9)'],
        ['EPS ($)', 1.52, 1.69, 1.85, 2.01, '=SUM(B10:E10)'],
    ]

    for row_num, row_data in enumerate(data, 4):
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num > 1:
                if row_num == 10:  # EPS row
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Segment Breakdown
    ws2 = wb.create_sheet("Segment Breakdown")

    ws2['A1'] = 'Revenue by Segment (Q3 2024)'
    ws2['A1'].font = Font(bold=True, size=14)

    seg_headers = ['Segment', 'Revenue ($M)', '% of Total', 'YoY Growth', 'Margin %']
    for col, header in enumerate(seg_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    segments = [
        ['Cloud Services', 1200, 0.50, 0.25, 0.72],
        ['Enterprise Software', 720, 0.30, 0.10, 0.68],
        ['Hardware', 360, 0.15, 0.05, 0.45],
        ['Professional Services', 120, 0.05, 0.08, 0.55],
    ]

    for row_num, seg in enumerate(segments, 4):
        for col_num, value in enumerate(seg, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 2:
                cell.number_format = '#,##0'
            elif col_num in [3, 4, 5]:
                cell.number_format = '0%'

    # Total
    total_row = 4 + len(segments)
    ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=f'=SUM(B4:B{total_row-1})').font = Font(bold=True)
    ws2.cell(row=total_row, column=2).number_format = '#,##0'
    ws2.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})').font = Font(bold=True)
    ws2.cell(row=total_row, column=3).number_format = '0%'

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    script_dir = Path(__file__).parent

    print("Creating sample test files...")
    print("=" * 40)

    # Create PDFs
    create_earnings_pdf(str(script_dir / "earnings_report_q3_2024.pdf"))
    create_pricing_pdf(str(script_dir / "product_pricing_2024.pdf"))

    # Create Excel files
    create_pricing_excel(str(script_dir / "pricing_calculator.xlsx"))
    create_earnings_excel(str(script_dir / "financial_summary_2024.xlsx"))

    print("=" * 40)
    print("Done!")
